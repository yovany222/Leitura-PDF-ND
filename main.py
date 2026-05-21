import os
import pandas as pd
from pypdf import PdfReader
from parsers import tag, tbg, nts
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# LER PDF
def ler_pdf(caminho):
    reader = PdfReader(caminho)
    texto = ""

    for page in reader.pages:
        texto += page.extract_text() or ""

    return texto


# IDENTIFICAR REMETENTE
def identificar_remetente(texto):
    if ("TRANSPORTADORA ASSOCIADA DE GÁS" in texto or "TRANSPORTADORA ASSOCIADA" in texto):
        return "TAG"
    elif ("NOVA TRANSPORTADORA DO SUDESTE" in texto or "NOVA TRANSPORTADORA" in texto):
        return "NTS"
    elif ("Transportadora Brasileira Gasoduto Bolívia Brasil" in texto or "Transportadora Brasileira" in texto):
        return "TBG"
    else:
        return None


# ESCOLHER PARSER
def escolher_parser(remetente):
    if remetente == "TAG":
        return tag.extrair_dados
    elif remetente == "TBG":
        return tbg.extrair_dados
    elif remetente == "NTS":
        return nts.extrair_dados
    else:
        return None


# SALVAR EXCEL
def salvar_excel(dados):
    arquivo = "dados_nd.xlsx"

    df = pd.DataFrame(dados)

    # ordem das colunas
    ordem = [
        "remetente",
        "cnpj eneva",
        "numero",
        "data_envio",
        "data_vencimento",
        "motivo",
        "valor",
        "arquivo"
    ]

    # garantir que só usa colunas existentes
    df = df[[col for col in ordem if col in df.columns]]

    # deixar colunas maiúsculas
    df.columns = [col.upper() for col in df.columns]

    df.to_excel(arquivo, index=False)

    wb = load_workbook(arquivo)
    ws = wb.active

    # FORMATO MONETÁRIO
    for col in ws.columns:
        if col[0].value == "VALOR":
            for cell in col[1:]:  # ignora cabeçalho
                if cell.value is not None:
                    cell.number_format = 'R$ #,##0.00'

    # CABEÇALHO
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        ws.row_dimensions[1].height = 30

    
    # CENTRALIZAR TUDO
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws["H"][1:]:
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # AUTOAJUSTE COLUNAS
    colunas_ignoradas = ["MOTIVO", "ARQUIVO"]

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        nome_coluna = col[0].value

        if nome_coluna in colunas_ignoradas:
            # largura fixa
            ws.column_dimensions[col_letter].width = 40
            continue

        max_length = 0

        for cell in col:
            if cell.value:
                tamanho = len(str(cell.value)) * 1.2
                if tamanho > max_length:
                    max_length = tamanho

        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
    
    # CORES
    cores = {
        "TAG": "38A0F6",
        "TBG": "090294",
        "NTS": "EB0B0B"
    }

    # REMETENTE
    for cell in ws["A"][1:]:
        valor = str(cell.value)

        cell.font = Font(bold=True, color="FFFFFF")
        if valor in cores:
            cell.fill = PatternFill(start_color=cores[valor],end_color=cores[valor],fill_type="solid")

    wb.save(arquivo)


# MAIN
def main():
    pasta = "PDFs"
    dados_extraidos = []

    for arquivo in os.listdir(pasta):
        if arquivo.endswith(".pdf"):
            caminho = os.path.join(pasta, arquivo)

            print(f"Processando: {arquivo}")

            texto = ler_pdf(caminho)

            remetente = identificar_remetente(texto)

            if not remetente:
                print("remetente não identificada\n")
                continue

            parser = escolher_parser(remetente)

            if not parser:
                print("Parser não encontrado\n")
                continue

            try:
                dados = parser(texto)
                dados["arquivo"] = arquivo

                dados_extraidos.append(dados)

            except Exception as e:
                print(f"Erro ao processar {arquivo}: {e}\n")

    # salvar no Excel
    if dados_extraidos:
        salvar_excel(dados_extraidos)
        print("✅ Planilha gerada: notas.xlsx")
    else:
        print("Nenhum dado extraído")


if __name__ == "__main__":
    main()